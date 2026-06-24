"""
add_new_restaurants.py
為沒有 WP_Post_ID 的餐廳自動補編號，上傳 placeholder 圖片，推送至 Firestore，並更新 xlsx。

執行方式：
  python "BYOB Finder Philadelphia/add_new_restaurants.py"

需要：pip install firebase-admin openpyxl
"""

from pathlib import Path
import firebase_admin
from firebase_admin import credentials, firestore, storage
import openpyxl

SCRIPT_DIR = Path(__file__).parent
BYOB_DIR   = SCRIPT_DIR.parent
SA_PATH    = BYOB_DIR / 'byob-app-5e4db-firebase-adminsdk-fbsvc-c8314f2fe3.json'
XLSX_PATH  = BYOB_DIR / 'philly_yelp_crawler' / 'data' / 'Philly BYOB Restaurant.xlsx'
BUCKET     = 'byob-app-5e4db.firebasestorage.app'
COLLECTION = 'restaurants'
MID_DIR    = BYOB_DIR / 'Mid'

# philly_restaurant_type token → image subfolder under Mid/
TYPE_FOLDER = {
    'italian':      'Italain',
    'mediterranean':'Mediterranean',
    'mexican':      'Mexican',
    'thai':         'Thai',
    'indian':       'Indian',
    'japanese':     'Sushi bar',
    'sushi':        'Sushi bar',
    'seafood':      'seafood',
    'ramen':        'Ramen',
    'pizza':        'pizza',
    'georgian':     'Georgian',
    'french':       'Fine dining',
    'american':     'Fine dining',
    'asian':        'Fine dining',
    'other':        'Placeholder',
}

def init_firebase():
    cred = credentials.Certificate(str(SA_PATH))
    firebase_admin.initialize_app(cred, {'storageBucket': BUCKET})
    return firestore.client(), storage.bucket()

def find_placeholder(rtype: str) -> Path | None:
    """Pick a .webp from the best-matching Mid subfolder."""
    for token in [t.strip().lower() for t in rtype.split(',')]:
        folder = TYPE_FOLDER.get(token)
        if folder:
            candidates = list((MID_DIR / folder / 'IMAGE_CONVERT').glob('*.webp'))
            if candidates:
                return candidates[0]
    # fallback
    fallback = list((MID_DIR / 'Placeholder').glob('*.webp'))
    return fallback[0] if fallback else None

def upload_image(doc_id: str, img_path: Path, bkt) -> str:
    blob = bkt.blob(f'images/restaurants/{doc_id}.webp')
    blob.upload_from_filename(str(img_path), content_type='image/webp')
    blob.make_public()
    return blob.public_url

def main():
    print('Initialising Firebase...')
    db, bkt = init_firebase()

    print('Reading xlsx...')
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    col = {h: i for i, h in enumerate(headers)}  # header → 0-based index

    # ── Find max existing ID and rows missing an ID ──────────────────────────
    max_id   = 0
    new_rows = []   # list of (excel_row_number_1based, row_values)

    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        wp_id = row[col['WP_Post_ID']].value
        name  = row[col['Name']].value
        if isinstance(wp_id, int) and wp_id > max_id:
            max_id = wp_id
        elif wp_id is None and name:
            new_rows.append((excel_row, row))

    print(f'Max existing WP_Post_ID : {max_id}')
    print(f'Restaurants without ID  : {len(new_rows)}')
    if not new_rows:
        print('Nothing to do.')
        return

    next_id = max_id + 1

    for excel_row, row in new_rows:
        doc_id = str(next_id)
        name   = row[col['Name']].value
        rtype  = str(row[col['philly_restaurant_type']].value or '')
        lat    = row[col['Latitude']].value
        lng    = row[col['Longitude']].value

        print(f'\n── {name}  →  ID {doc_id} ──')

        # Upload placeholder image
        img_path = find_placeholder(rtype)
        if img_path:
            print(f'  Uploading: {img_path.name}')
            cover_url = upload_image(doc_id, img_path, bkt)
            print(f'  URL: {cover_url}')
        else:
            cover_url = ''
            print('  WARNING: no placeholder image found, cover_image_url will be empty')

        # Build Firestore document
        lat_f = float(lat) if lat is not None else 0.0
        lng_f = float(lng) if lng is not None else 0.0

        amount_raw = row[col['corkage_fee_amount']].value
        doc_data = {
            'Name':                              name or '',
            'Add':                               row[col['Add']].value or '',
            'Phone':                             str(row[col['Phone']].value or ''),
            'philly_restaurant_type':            rtype,
            'philly_restaurant_type_other_note': row[col['philly_restaurant_type_other_note']].value or '',
            'philly_corkage_fee':                row[col['philly_corkage_fee']].value or '',
            'corkage_fee_amount':                float(amount_raw) if amount_raw else 0.0,
            'cover_image_url':                   cover_url,
            'Latitude':                          lat_f,
            'Longitude':                         lng_f,
            'location':                          firestore.GeoPoint(lat_f, lng_f),
        }

        db.collection(COLLECTION).document(doc_id).set(doc_data)
        print(f'  Pushed to Firestore: {COLLECTION}/{doc_id}')

        # Write new ID + cover URL back into xlsx
        ws.cell(row=excel_row, column=col['WP_Post_ID'] + 1).value = next_id
        ws.cell(row=excel_row, column=col['cover_image_url'] + 1).value = cover_url

        next_id += 1

    wb.save(str(XLSX_PATH))
    print(f'\n✅  xlsx saved: {XLSX_PATH}')
    print(f'✅  Done — added {len(new_rows)} new restaurant(s).')

if __name__ == '__main__':
    main()
