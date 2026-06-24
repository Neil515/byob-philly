"""
migrate_restaurant_types.py
修正 Firestore 和 xlsx 的 philly_restaurant_type 欄位：

1. other + note 含 Pizza  → 把 other token 換成 pizza
2. other + note 含 Sushi  → 把 other token 換成 sushi
3. other + note 含 Ramen  → 把 other token 換成 ramen
4. ID=1790 Chinese       → type=other, note=Chinese（讓 Other chip 能找到）
5. 清除 type 欄位裡的 non-breaking space (\\xa0)

執行方式：
  python "BYOB Finder Philadelphia/migrate_restaurant_types.py"

需要：pip install firebase-admin openpyxl
"""

from pathlib import Path
import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl

SCRIPT_DIR = Path(__file__).parent
BYOB_DIR   = SCRIPT_DIR.parent
SA_PATH    = BYOB_DIR / 'byob-app-5e4db-firebase-adminsdk-fbsvc-c8314f2fe3.json'
XLSX_PATH  = BYOB_DIR / 'philly_yelp_crawler' / 'data' / 'Philly BYOB Restaurant.xlsx'
COLLECTION = 'restaurants'

# note keyword → new type token (priority: first match wins)
NOTE_TO_TYPE = [
    ('ramen',  'ramen'),
    ('pizza',  'pizza'),
    ('sushi',  'sushi'),
]

def resolve_type(current_type: str, note: str) -> str:
    """Replace 'other' token with specific type if note matches a known keyword."""
    note_lower = note.lower()
    for keyword, new_token in NOTE_TO_TYPE:
        if keyword in note_lower:
            # Replace 'other' token with new_token (case-insensitive match)
            tokens = [t.strip() for t in current_type.split(',')]
            tokens = [new_token if t.lower() == 'other' else t for t in tokens]
            return ', '.join(tokens)
    return current_type

def normalize(s: str) -> str:
    """Remove non-breaking spaces and trim."""
    return s.replace('\xa0', '').replace('  ', ' ').strip().strip(',').strip()

def main():
    print('Initialising Firebase...')
    cred = credentials.Certificate(str(SA_PATH))
    firebase_admin.initialize_app(cred)
    db = firestore.client()

    print('Reading xlsx...')
    wb = openpyxl.load_workbook(str(XLSX_PATH))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    print('Fetching Firestore documents...')
    docs = list(db.collection(COLLECTION).stream())
    print(f'Found {len(docs)} documents\n')

    changes = []

    for doc in docs:
        data    = doc.to_dict()
        doc_id  = doc.id
        name    = data.get('Name', doc_id)
        rtype   = normalize(str(data.get('philly_restaurant_type', '')))
        note    = str(data.get('philly_restaurant_type_other_note', '') or '')
        updates = {}

        # ── Rule 4: Chinese standalone → other + note ────────────────────
        if doc_id == '1790':
            if rtype == 'Chinese':
                updates['philly_restaurant_type']            = 'other'
                updates['philly_restaurant_type_other_note'] = 'Chinese'
                print(f'  [1790] {name}: type Chinese → other, note → Chinese')

        # ── Rule 1-3: other + pizza/sushi/ramen note ─────────────────────
        elif 'other' in rtype.lower() and note:
            new_type = resolve_type(rtype, note)
            if new_type != rtype:
                updates['philly_restaurant_type'] = new_type
                print(f'  [{doc_id}] {name}')
                print(f'    type: {rtype!r} → {new_type!r}  (note: {note!r})')

        # ── Rule 5: normalize whitespace for all docs ────────────────────
        normalized = normalize(rtype)
        if normalized != data.get('philly_restaurant_type', ''):
            updates['philly_restaurant_type'] = updates.get(
                'philly_restaurant_type', normalized)

        if updates:
            db.collection(COLLECTION).document(doc_id).update(updates)
            changes.append((doc_id, name, updates))

    print(f'\nFirestore: {len(changes)} document(s) updated.')

    # ── Update xlsx ───────────────────────────────────────────────────────
    xlsx_updates = 0
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        wp_id = str(row[col['WP_Post_ID']].value or '')
        for doc_id, name, updates in changes:
            if wp_id == doc_id:
                if 'philly_restaurant_type' in updates:
                    ws.cell(row=excel_row,
                            column=col['philly_restaurant_type'] + 1
                            ).value = updates['philly_restaurant_type']
                if 'philly_restaurant_type_other_note' in updates:
                    ws.cell(row=excel_row,
                            column=col['philly_restaurant_type_other_note'] + 1
                            ).value = updates['philly_restaurant_type_other_note']
                xlsx_updates += 1

    wb.save(str(XLSX_PATH))
    print(f'xlsx: {xlsx_updates} row(s) updated.')
    print(f'\n✅  Done.')

if __name__ == '__main__':
    main()
